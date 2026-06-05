"""XLSX writer using openpyxl with a touch of styling."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.exports.builder import ExportSheet

HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="0F1F1A")
HEADER_FILL = PatternFill("solid", fgColor="E6FFF1")
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="DDDDDD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def render_xlsx(sheets: list[ExportSheet], *, title: str | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.name[:31])
        for ci, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BOX
            cell.alignment = Alignment(vertical="center")
        for ri, row in enumerate(sheet.rows, start=2):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = BODY_FONT
                cell.border = BOX
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.alignment = Alignment(horizontal="right")
        ws.freeze_panes = "A2"
        for ci, col in enumerate(sheet.columns, start=1):
            letter = get_column_letter(ci)
            longest = max([len(str(col))] + [len(str(r[ci-1])) for r in sheet.rows if ci - 1 < len(r)])
            ws.column_dimensions[letter].width = min(60, max(10, longest + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
